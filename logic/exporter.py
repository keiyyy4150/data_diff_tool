import os
import pandas as pd
import copy
from datetime import datetime
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def export_diff(shared):
    file1 = shared["file1_path"].get()
    df1 = shared["df1"].copy().astype(str)
    df2 = shared["df2"].copy().astype(str)
    key1 = shared["key_col1"].get()
    key2 = shared["key_col2"].get()
    col_mapping = shared["col_mapping"]

    output_format = choose_format()
    if output_format not in ["xlsx", "csv"]:
        raise Exception("出力形式が選択されていません")

    # 差分処理（本来はdiff_logicに分けてもよい）
    df1_cols = list(df1.columns)
    df2_cols = list(df2.columns)
    added_columns = [col for col in df2_cols if col not in col_mapping.values()]
    renamed_columns = [(old, new) for old, new in col_mapping.items() if old != new]
    deleted_columns = [col for col in df1_cols if col not in col_mapping]

    output_cols = [col for col in df1_cols if col not in deleted_columns] + added_columns
    display_cols = []
    for col in output_cols:
        rename_match = next(((old, new) for old, new in renamed_columns if old == col), None)
        if rename_match:
            display_cols.append(f"変更: {rename_match[1]} ←→ {rename_match[0]}")
        elif col in added_columns:
            display_cols.append(f"追加: {col}")
        else:
            display_cols.append(col)

    merged = pd.merge(df1, df2, left_on=key1, right_on=key2, how='outer', indicator=True, suffixes=('_old', '_new'))
    output_data = [display_cols]

    for _, row in merged.iterrows():
        row_data = []
        for col in output_cols:
            if col in added_columns:
                val = row.get(col, "")
                row_data.append(f"追加: {val}" if row['_merge'] != 'left_only' else "")
            else:
                col2 = col_mapping.get(col, "")
                val1 = row.get(f"{col}_old", row.get(col, ""))
                val2 = row.get(f"{col2}_new", row.get(col2, ""))
                if row['_merge'] == 'both':
                    if val1 == val2:
                        row_data.append(val1)
                    else:
                        row_data.append(f"変更: {val1} ←→ {val2}")
                elif row['_merge'] == 'left_only':
                    row_data.append(f"削除: {val1}")
                elif row['_merge'] == 'right_only':
                    row_data.append(f"追加: {val2}")
        output_data.append(row_data)

    ext = f".{output_format}"
    output_path = os.path.join(os.path.dirname(file1), f"差分結果{ext}")

    if output_format == "csv":
        pd.DataFrame(output_data).to_csv(output_path, index=False, header=False, encoding='utf-8-sig')
    else:
        wb = Workbook()
        ws_diff = wb.active
        ws_diff.title = "差分結果"
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in output_data:
            ws_diff.append(row)
        for row in ws_diff.iter_rows():
            for cell in row:
                cell.number_format = '@'
                if cell.value and any(x in str(cell.value) for x in ['←→', '削除', '追加']):
                    cell.fill = copy.copy(yellow)
        # 元データも出力
        for df, label in [(shared["df1"], os.path.splitext(os.path.basename(file1))[0]),
                          (shared["df2"], os.path.splitext(os.path.basename(shared["file2_path"].get()))[0])]:
            ws = wb.create_sheet(title=label)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        wb.save(output_path)

def choose_format():
    from tkinter import Toplevel, Label, Radiobutton, StringVar, Button
    top = Toplevel()
    top.title("出力形式を選択")
    var = StringVar()
    Label(top, text="出力形式を選んでください:").pack(pady=10)
    Radiobutton(top, text=".xlsx (Excel形式)", variable=var, value="xlsx").pack()
    Radiobutton(top, text=".csv (UTF-8)", variable=var, value="csv").pack()
    Button(top, text="OK", command=top.destroy).pack(pady=10)
    top.grab_set()
    top.wait_window()
    return var.get()