import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import copy
from datetime import datetime

class ExcelDiffApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel/CSV差分比較ツール")

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.start_row1 = tk.StringVar()
        self.start_row2 = tk.StringVar()
        self.col_mapping = {}
        self.key_col1 = tk.StringVar()
        self.key_col2 = tk.StringVar()
        self.file_ext = None

        self.init_frame1()

    def clear_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def select_file(self, var):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv")])
        if path:
            var.set(path)

    def init_frame1(self):
        self.clear_frame()
        frame = tk.Frame(self.root)
        frame.pack(padx=10, pady=10)

        tk.Label(frame, text="比較元ファイル:").grid(row=0, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.file1_path, width=50).grid(row=0, column=1, sticky="w")
        tk.Button(frame, text="参照", command=lambda: self.select_file(self.file1_path)).grid(row=0, column=2)

        tk.Label(frame, text="開始行（比較元）:").grid(row=1, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.start_row1).grid(row=1, column=1, sticky="w")

        tk.Label(frame, text="比較先ファイル:").grid(row=2, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.file2_path, width=50).grid(row=2, column=1, sticky="w")
        tk.Button(frame, text="参照", command=lambda: self.select_file(self.file2_path)).grid(row=2, column=2)

        tk.Label(frame, text="開始行（比較先）:").grid(row=3, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.start_row2).grid(row=3, column=1, sticky="w")

        tk.Button(frame, text="終了", command=self.root.quit).grid(row=4, column=1, sticky="e", pady=10)
        tk.Button(frame, text="次へ", command=self.go_to_key_selection).grid(row=4, column=2, sticky="w")

    def go_to_key_selection(self):
        if not all([self.file1_path.get(), self.file2_path.get(), self.start_row1.get(), self.start_row2.get()]):
            messagebox.showerror("エラー", "全ての項目を入力してください")
            return

        ext1 = os.path.splitext(self.file1_path.get())[1].lower()
        ext2 = os.path.splitext(self.file2_path.get())[1].lower()

        if ext1 != ext2:
            messagebox.showerror("拡張子不一致", "比較元と比較先のファイル形式（拡張子）が異なります。両方CSVまたは両方Excelファイルにしてください。")
            return

        self.file_ext = ext1

        try:
            if ext1 == '.csv':
                self.df1 = pd.read_csv(self.file1_path.get(), header=int(self.start_row1.get()) - 1, dtype=str)
                self.df2 = pd.read_csv(self.file2_path.get(), header=int(self.start_row2.get()) - 1, dtype=str)
            else:
                self.df1 = pd.read_excel(self.file1_path.get(), header=int(self.start_row1.get()) - 1)
                self.df2 = pd.read_excel(self.file2_path.get(), header=int(self.start_row2.get()) - 1)
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e))
            return

        self.clear_frame()
        frame = tk.Frame(self.root)
        frame.pack(padx=10, pady=10)

        tk.Label(frame, text="キー列（比較元側）:").grid(row=0, column=0, sticky="w")
        ttk.Combobox(frame, textvariable=self.key_col1, values=list(self.df1.columns), state="readonly").grid(row=0, column=1, sticky="w")

        tk.Label(frame, text="キー列（比較先側）:").grid(row=1, column=0, sticky="w")
        ttk.Combobox(frame, textvariable=self.key_col2, values=list(self.df2.columns), state="readonly").grid(row=1, column=1, sticky="w")

        tk.Button(frame, text="戻る", command=self.init_frame1).grid(row=2, column=0, pady=10, sticky="w")
        tk.Button(frame, text="終了", command=self.root.quit).grid(row=2, column=1, pady=10, sticky="e")
        tk.Button(frame, text="次へ", command=self.go_to_mapping).grid(row=2, column=2, pady=10, sticky="w")

    def go_to_mapping(self):
        self.clear_frame()
        frame = tk.Frame(self.root)
        frame.pack(padx=10, pady=10)

        tk.Label(frame, text="比較元カラム", width=30).grid(row=0, column=0, sticky="w")
        tk.Label(frame, text="比較先カラム", width=30).grid(row=0, column=1, sticky="w")

        self.combos = {}
        for i, col1 in enumerate(self.df1.columns):
            tk.Label(frame, text=col1).grid(row=i + 1, column=0, sticky="w")
            combo = ttk.Combobox(frame, values=list(self.df2.columns), state="readonly")
            combo.grid(row=i + 1, column=1, sticky="w")
            if col1 in self.df2.columns:
                combo.set(col1)
            self.combos[col1] = combo

        button_row = i + 2
        tk.Button(frame, text="戻る", command=self.go_to_key_selection).grid(row=button_row, column=0, pady=10, sticky="w")
        tk.Button(frame, text="終了", command=self.root.quit).grid(row=button_row, column=1, pady=10, sticky="e", padx=(0, 80))
        tk.Button(frame, text="実行する", command=self.execute_diff).grid(row=button_row, column=1, pady=10, sticky="e")

    def execute_diff(self):
        try:
            for k, v in self.combos.items():
                self.col_mapping[k] = v.get()

            output_format = tk.StringVar()
            def select_format():
                popup = tk.Toplevel()
                popup.title("出力形式を選択")
                tk.Label(popup, text="出力形式を選んでください:").pack(pady=10)
                tk.Radiobutton(popup, text=".xlsx (Excel形式)", variable=output_format, value="xlsx").pack()
                tk.Radiobutton(popup, text=".csv (UTF-8)", variable=output_format, value="csv").pack()
                tk.Button(popup, text="OK", command=popup.destroy).pack(pady=10)
                popup.grab_set()
                self.root.wait_window(popup)

            select_format()
            if output_format.get() not in ["xlsx", "csv"]:
                messagebox.showerror("キャンセル", "出力形式が選択されていません。")
                return

            df1 = self.df1.copy()
            df2 = self.df2.copy()
            key1 = self.key_col1.get()
            key2 = self.key_col2.get()

            df1 = df1.astype(str)
            df2 = df2.astype(str)

            df1_cols = list(df1.columns)
            df2_cols = list(df2.columns)
            added_columns = [col for col in df2_cols if col not in self.col_mapping.values()]
            renamed_columns = [(old, new) for old, new in self.col_mapping.items() if old != new]
            deleted_columns = [col for col in df1_cols if col not in self.col_mapping]

            output_cols = [col for col in df1_cols if col not in deleted_columns] + added_columns
            display_cols = []
            for col in output_cols:
                rename_match = next(((old, new) for old, new in renamed_columns if old == col), None)
                if rename_match:
                    display_cols.append(f"変更: {rename_match[1]} ←→ {rename_match[0]}")
                elif col in added_columns:
                    display_cols.append(f"追加: {col}")
                else:
                    display_cols.append(col)

            merged = pd.merge(df1, df2, left_on=key1, right_on=key2, how='outer', indicator=True, suffixes=('_old', '_new'))

            output_data = [display_cols]

            for _, row in merged.iterrows():
                output_row = []
                for col in output_cols:
                    if col in added_columns:
                        val = row.get(col, "")
                        output_row.append(f"追加: {val}" if row['_merge'] != 'left_only' else "")
                    else:
                        col2 = self.col_mapping.get(col, "")
                        val1 = row.get(f"{col}_old", row.get(col, ""))
                        val2 = row.get(f"{col2}_new", row.get(col2, ""))
                        if row['_merge'] == 'both':
                            if pd.isna(val1) and pd.isna(val2):
                                output_row.append("")
                            elif val1 == val2:
                                output_row.append(val1)
                            else:
                                output_row.append(f"変更: {val1} ←→ {val2}")
                        elif row['_merge'] == 'left_only':
                            output_row.append(f"削除: {val1}")
                        elif row['_merge'] == 'right_only':
                            output_row.append(f"追加: {val2}")
                output_data.append(output_row)

            ext = f".{output_format.get()}"
            output_path = os.path.join(os.path.dirname(self.file1_path.get()), f"差分結果{ext}")

            if output_format.get() == "csv":
                pd.DataFrame(output_data).to_csv(output_path, index=False, header=False, encoding='utf-8-sig')
            else:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                wb = Workbook()
                ws_diff = wb.active
                ws_diff.title = "差分結果"
                yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                for row in output_data:
                    ws_diff.append(row)

                for row in ws_diff.iter_rows(min_row=1, max_row=ws_diff.max_row):
                    for cell in row:
                        cell.number_format = '@'
                        if cell.value and any(k in str(cell.value) for k in ['←→', '削除', '追加', '列追加', '列名変更']):
                            cell.fill = copy.copy(yellow)

                def add_sheet(wb, df, title):
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    ws = wb.create_sheet(title=title)
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
                        for j, cell in enumerate(row):
                            value = df.iloc[i - 1, j]
                            if pd.isna(value):
                                continue
                            if isinstance(value, datetime):
                                cell.number_format = 'yyyy/m/d'
                            elif isinstance(value, (int, float)):
                                cell.number_format = 'General'
                            else:
                                cell.number_format = '@'

                add_sheet(wb, self.df1, os.path.splitext(os.path.basename(self.file1_path.get()))[0])
                add_sheet(wb, self.df2, os.path.splitext(os.path.basename(self.file2_path.get()))[0])
                wb.save(output_path)

            messagebox.showinfo("完了", f"差分ファイルを出力しました：\n{output_path}")
            self.root.destroy()

        except Exception as e:
            messagebox.showerror("比較エラー", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelDiffApp(root)
    root.mainloop()